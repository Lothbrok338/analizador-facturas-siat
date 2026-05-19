import streamlit as st
import pandas as pd
from urllib.parse import urlparse, parse_qs
from io import BytesIO, StringIO
import os
import re
import json
import gspread
from google.oauth2.service_account import Credentials
from gspread_dataframe import set_with_dataframe, get_as_dataframe
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Gestión Contable | UNIVALLE", page_icon="🎓", layout="wide")

# --- COLECION DE COLUMNAS CONFIGURADAS ---
ORDEN_COLUMNAS_SISTEMA = [
    "Fecha", "Razón Social", "NIT", "Nro Factura", "CUF / Autorización",
    "IMPORTE TOTAL COMPRA", "IMPORTE ICE", "TASAS", "SUBTOTAL", 
    "DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA", "IMPORTE BASE CF", "CREDITO FISCAL"
]

# --- CONEXIÓN A GOOGLE SHEETS ---
@st.cache_resource
def init_connection():
    creds_dict = json.loads(st.secrets["google_credentials_json"])
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    client = gspread.authorize(creds)
    return client.open_by_url(st.secrets["spreadsheet_url"])

# --- FUNCIONES DE BASE DE DATOS EN LA NUBE ---
@st.cache_data(ttl=600)
def cargar_historico():
    try:
        sheet = init_connection()
        ws = sheet.worksheet("HISTORICO_FACTURAS")
        df = get_as_dataframe(ws).dropna(how='all').dropna(axis=1, how='all')
        if df.empty or 'CUF / Autorización' not in df.columns:
            return pd.DataFrame(columns=ORDEN_COLUMNAS_SISTEMA)
        return df
    except Exception as e:
        return pd.DataFrame(columns=ORDEN_COLUMNAS_SISTEMA)

def guardar_historico(df_nuevo):
    sheet = init_connection()
    ws = sheet.worksheet("HISTORICO_FACTURAS")
    df_actual = cargar_historico()
    df_final = pd.concat([df_actual, df_nuevo], ignore_index=True)
    
    df_final = df_final[ORDEN_COLUMNAS_SISTEMA]
    
    ws.clear()
    set_with_dataframe(ws, df_final)
    cargar_historico.clear()

def eliminar_de_historico_nube(cuf_eliminar):
    try:
        sheet = init_connection()
        ws = sheet.worksheet("HISTORICO_FACTURAS")
        celda = ws.find(str(cuf_eliminar).strip())
        if celda:
            ws.delete_rows(celda.row)
            cargar_historico.clear()
            return True
    except:
        pass
    return False

@st.cache_data(ttl=600)
def cargar_siat_maestro():
    try:
        sheet = init_connection()
        ws = sheet.worksheet("SIAT_MAESTRO")
        df = get_as_dataframe(ws).dropna(how='all').dropna(axis=1, how='all')
        if df.empty:
            return None
        return df
    except:
        return None

def guardar_siat_maestro(df_nuevo):
    sheet = init_connection()
    ws = sheet.worksheet("SIAT_MAESTRO")
    df_actual = cargar_siat_maestro()
    
    if df_actual is not None and not df_actual.empty:
        df_combinado = pd.concat([df_actual, df_nuevo], ignore_index=True)
        df_combinado = df_combinado.drop_duplicates(subset=['CODIGO DE AUTORIZACION'], keep='last')
    else:
        df_combinado = df_nuevo
        
    ws.clear()
    set_with_dataframe(ws, df_combinado)
    cargar_siat_maestro.clear()

# --- MOTOR DE ESTILIZACIÓN MINIMALISTA EXCEL ---
def estilizar_hoja_excel(ws):
    # Forzar visibilidad de líneas de cuadrícula internas
    ws.views.sheetView[0].showGridLines = True
    
    # Paleta y Tipografía Minimalista
    fuente_encabezado = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    relleno_encabezado = PatternFill(start_color='741B28', end_color='741B28', fill_type='solid') # Vino Institucional
    fuente_cuerpo = Font(name='Segoe UI', size=10, bold=False, color='1A1A1A')
    relleno_cebra = PatternFill(start_color='FDFDF9', end_color='FDFDF9', fill_type='solid') # Crema Marfil tenue
    
    borde_sutil = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # 1. Estilizar fila de encabezados
    for cell in ws[1]:
        cell.font = fuente_encabezado
        cell.fill = relleno_encabezado
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borde_sutil
    ws.row_dimensions[1].height = 28
    
    # 2. Estilizar filas de datos
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
        ws.row_dimensions[row_idx].height = 20
        es_par = (row_idx % 2 == 0)
        
        for cell in row:
            cell.font = fuente_cuerpo
            cell.border = borde_sutil
            if es_par:
                cell.fill = relleno_cebra
                
            # Formatear y alinear según el tipo de dato
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
            elif cell.value and (any(x in str(cell.value) for x in ['-', '/']) or str(cell.value).isdigit() and len(str(cell.value)) > 5):
                # Centrar fechas, NITs, números de factura largos
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
    # 3. Auto-ajustar el ancho de las columnas dinámicamente
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

# --- ESTILOS CSS PROFESIONALES APP ---
st.markdown("""
<style>
    .stApp { background-color: #fdf5e6; }
    [data-testid="stSidebar"] { background-color: #741b28 !important; border-right: 2px solid #b8860b; }
    [data-testid="stSidebar"] div, [data-testid="stSidebar"] span, 
    [data-testid="stSidebar"] label, [data-testid="stSidebar"] p { color: #ffffff !important; font-weight: 500 !important; }
    [data-testid="stFileUploader"] section { background-color: #1a1a1a !important; border: 1px dashed #b8860b !important; border-radius: 8px !important; }
    [data-testid="stFileUploaderFileName"], [data-testid="stFileUploaderFileData"], 
    [data-testid="stFileUploader"] small { color: #b8860b !important; opacity: 1 !important; }
    [data-testid="stFileUploader"] button { background-color: #741b28 !important; color: white !important; border: 1px solid #b8860b !important; }
    .stButton > button { border-radius: 4px; font-weight: 600; text-transform: uppercase; transition: all 0.3s ease; }
    div.stButton > button:first-child:not([kind="primary"]) { background-color: #741b28 !important; color: #ffffff !important; border: 1px solid #b8860b !important; }
    .stButton > button[kind="primary"] { background-color: #741b28 !important; color: #ffffff !important; border: 1px solid #b8860b !important; height: 3em; }
    h1, h2, h3 { color: #741b28; font-family: 'Times New Roman', serif; }
    .factura-card { background-color: #ffffff; padding: 15px; border-left: 5px solid #741b28; border-radius: 4px; margin-bottom: 12px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }
    .cuf-text { color: #b8860b; font-family: monospace; font-weight: bold; }
    .alerta-duplicado { color: #d32f2f; font-weight: bold; background-color: #ffebee; padding: 10px; border-radius: 5px; margin-bottom: 10px; border-left: 4px solid #d32f2f;}
    .bandeja-box { background-color: #e8f5e9; padding: 15px; border-radius: 8px; border: 1px solid #4caf50; margin-bottom: 20px; }
</style>
""", unsafe_allow_html=True)

# --- LÓGICA DE SESIÓN ---
if 'registros_sesion' not in st.session_state:
    st.session_state.registros_sesion = []
if 'registros_pendientes' not in st.session_state:
    st.session_state.registros_pendientes = []

# --- PANEL LATERAL ---
with st.sidebar:
    logo_path = "UNIVALLE LOGO.webp"
    if os.path.exists(logo_path):
        st.image(logo_path, use_container_width=True)
    else:
        st.markdown("<h2 style='color:white; text-align:center;'>UNIVALLE</h2>", unsafe_allow_html=True)
    
    st.markdown("<h4 style='text-align: center;'>INSTRUMENTO DE CONTROL CONTABLE</h4>", unsafe_allow_html=True)
    st.divider()
    
    archivo_csv = st.file_uploader("Vincular Base SIAT Diaria (.csv)", type=['csv'])
    
    if archivo_csv:
        try:
            content = archivo_csv.read()
            try:
                decoded_content = content.decode('utf-8')
            except UnicodeDecodeError:
                decoded_content = content.decode('latin1')
            
            df_diario = pd.read_csv(StringIO(decoded_content), sep=',', on_bad_lines='skip')
            df_diario.columns = [c.strip() for c in df_diario.columns]
            df_diario = df_diario.map(lambda x: x.strip() if isinstance(x, str) else x)
            
            with st.spinner("Guardando base en Google Sheets..."):
                guardar_siat_maestro(df_diario)
            st.success("Base diaria respaldada con éxito")
        except Exception as e:
            st.error(f"Error en la lectura: {e}")
    
    st.divider()
    
    df_historico_actual = cargar_historico()
    df_siat_actual = cargar_siat_maestro()
    
    st.write("☁️ **Estadísticas en la Nube:**")
    st.write(f"- Base Maestra: {len(df_siat_actual) if df_siat_actual is not None else 0} registros")
    st.write(f"- Facturas Procesadas: {len(df_historico_actual)}")
    
    st.divider()
    if st.button("🔄 Limpiar Pantalla", use_container_width=True):
        st.session_state.registros_sesion = []
        st.session_state.registros_pendientes = []
        st.rerun()

# --- INTERFAZ PRINCIPAL ---
st.title("UNIVERSIDAD DEL VALLE S.A.")
st.subheader("Módulo Centralizado de Procesamiento de Datos Fiscales")
st.divider()

base_siat = cargar_siat_maestro()

if base_siat is not None:
    tab1, tab2 = st.tabs(["📥 Consolidación por Enlaces (QR)", "🔍 Búsqueda y Registro Manual"])
    
    # PESTAÑA 1: PROCESAMIENTO POR QR / URL
    with tab1:
        urls_raw = st.text_area("Depósito de URLs para procesamiento masivo:", height=150, placeholder="Pega los enlaces aquí...", key="txt_urls")
        
        if st.button("🚀 PROCESAR ENLACES", type="primary", use_container_width=True, key="btn_qr"):
            links = re.findall(r'https?://[^\s]+?(?=https?://|$)', urls_raw)
            historico_db = cargar_historico()
            cufs_historicos = historico_db['CUF / Autorización'].tolist() if not historico_db.empty else []
            
            agregados = 0
            duplicados = 0
            
            with st.spinner("Buscando enlaces en la base maestra..."):
                for link in links:
                    try:
                        link_clean = link.strip().rstrip(',').rstrip(';')
                        params = parse_qs(urlparse(link_clean).query)
                        cuf_extraido = params.get('cuf', [''])[0].strip()
                        
                        if not cuf_extraido:
                            continue

                        if cuf_extraido in cufs_historicos or \
                           any(d['CUF / Autorización'] == cuf_extraido for d in st.session_state.registros_sesion) or \
                           any(d['CUF / Autorización'] == cuf_extraido for d in st.session_state.registros_pendientes):
                            duplicados += 1
                            continue

                        match = base_siat[base_siat['CODIGO DE AUTORIZACION'] == cuf_extraido]
                        
                        if not match.empty:
                            item = match.iloc[0]
                            rs_raw = str(item['RAZON SOCIAL PROVEEDOR'])
                            try:
                                razon_social = rs_raw.encode('latin1').decode('utf-8') if "Ã" in rs_raw else rs_raw
                            except:
                                razon_social = rs_raw
                            
                            nuevo_registro = {
                                "Fecha": item.get('FECHA DE FACTURA/DUI/DIM', ''),
                                "Razón Social": razon_social,
                                "NIT": item.get('NIT PROVEEDOR', ''),
                                "Nro Factura": item.get('NUMERO FACTURA', ''),
                                "CUF / Autorización": cuf_extraido,
                                "IMPORTE TOTAL COMPRA": item.get('IMPORTE TOTAL COMPRA', 0),
                                "IMPORTE ICE": item.get('IMPORTE ICE', 0),
                                "TASAS": item.get('TASAS', 0),
                                "SUBTOTAL": item.get('SUBTOTAL', 0),
                                "DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA": item.get('DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA', 0),
                                "IMPORTE BASE CF": item.get('IMPORTE BASE CF', 0),
                                "CREDITO FISCAL": item.get('CREDITO FISCAL', 0)
                            }
                            st.session_state.registros_pendientes.append(nuevo_registro)
                            agregados += 1
                    except:
                        continue
                
                if agregados > 0:
                    st.success(f"Búsqueda exitosa: Se enviaron {agregados} registros a la Bandeja de Revisión abajo.")
                if duplicados > 0:
                    st.markdown(f"<div class='alerta-duplicado'>⚠️ SE IGNORARON {duplicados} ENLACES (Ya registrados o en revisión).</div>", unsafe_allow_html=True)

    # PESTAÑA 2: MOTOR DE BÚSQUEDA MANUAL
    with tab2:
        st.markdown("### Búsqueda de Comprobante en Base Maestra")
        col_f, col_n = st.columns(2)
        with col_f:
            nro_factura_in = st.text_input("Número de Factura (Opcional):", placeholder="Ej. 4032", key="input_nro_fac")
        with col_n:
            nit_in = st.text_input("NIT del Proveedor (Opcional):", placeholder="Ej. 1020304029", key="input_nit")
            
        if st.button("🔍 LOCALIZAR REGISTROS", type="primary", use_container_width=True, key="btn_manual"):
            nro_val = nro_factura_in.strip()
            nit_val = nit_in.strip()
            
            if not nro_val and not nit_val:
                st.warning("⚠️ Por favor, introduzca al menos un criterio para iniciar la búsqueda.")
            else:
                base_siat['NUMERO FACTURA_STR'] = base_siat['NUMERO FACTURA'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                base_siat['NIT_STR'] = base_siat['NIT PROVEEDOR'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                
                if nro_val and nit_val:
                    match_manual = base_siat[(base_siat['NUMERO FACTURA_STR'] == nro_val) & (base_siat['NIT_STR'] == nit_val)]
                elif nro_val:
                    match_manual = base_siat[base_siat['NUMERO FACTURA_STR'] == nro_val]
                else:
                    match_manual = base_siat[base_siat['NIT_STR'] == nit_val]
                
                if match_manual.empty:
                    st.error("No se localizó ningún registro coincidente.")
                else:
                    historico_db = cargar_historico()
                    cufs_historicos = historico_db['CUF / Autorización'].tolist() if not historico_db.empty else []
                    
                    agregados_m = 0
                    duplicados_m = 0
                    
                    for idx, item in match_manual.iterrows():
                        cuf_extraido = str(item['CODIGO DE AUTORIZACION']).strip()
                        
                        if cuf_extraido in cufs_historicos or \
                           any(d['CUF / Autorización'] == cuf_extraido for d in st.session_state.registros_sesion) or \
                           any(d['CUF / Autorización'] == cuf_extraido for d in st.session_state.registros_pendientes):
                            duplicados_m += 1
                            continue
                            
                        rs_raw = str(item['RAZON SOCIAL PROVEEDOR'])
                        try:
                            razon_social = rs_raw.encode('latin1').decode('utf-8') if "Ã" in rs_raw else rs_raw
                        except:
                            razon_social = rs_raw
                            
                        nuevo_registro = {
                            "Fecha": item.get('FECHA DE FACTURA/DUI/DIM', ''),
                            "Razón Social": razon_social,
                            "NIT": item.get('NIT PROVEEDOR', ''),
                            "Nro Factura": item.get('NUMERO FACTURA', ''),
                            "CUF / Autorización": cuf_extraido,
                            "IMPORTE TOTAL COMPRA": item.get('IMPORTE TOTAL COMPRA', 0),
                            "IMPORTE ICE": item.get('IMPORTE ICE', 0),
                            "TASAS": item.get('TASAS', 0),
                            "SUBTOTAL": item.get('SUBTOTAL', 0),
                            "DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA": item.get('DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA', 0),
                            "IMPORTE BASE CF": item.get('IMPORTE BASE CF', 0),
                            "CREDITO FISCAL": item.get('CREDITO FISCAL', 0)
                        }
                        st.session_state.registros_pendientes.append(nuevo_registro)
                        agregados_m += 1
                    
                    if agregados_m > 0:
                        st.success(f"✅ Se enviaron {agregados_m} registros coincidentes a la Bandeja de Revisión abajo.")
                    if duplicados_m > 0:
                        st.markdown(f"<div class='alerta-duplicado'>⚠️ CONTROL DE DUPLICADOS: {duplicados_m} registros ya existían en el sistema.</div>", unsafe_allow_html=True)

# --- BANDEJA DE REVISIÓN (STAGING AREA) ---
if st.session_state.registros_pendientes:
    st.divider()
    st.markdown("<div class='bandeja-box'><h3>📥 Bandeja de Revisión</h3><p>Verifica los datos obtenidos. Las columnas se han ordenado para que veas el Monto inmediatamente sin necesidad de desplazar la barra.</p></div>", unsafe_allow_html=True)
    
    df_pendientes = pd.DataFrame(st.session_state.registros_pendientes)
    df_pendientes.insert(0, "Guardar", True)
    
    columnas_orden_tabla = [
        "Guardar", "Fecha", "Razón Social", "NIT", "Nro Factura", "IMPORTE BASE CF", "CUF / Autorización",
        "IMPORTE TOTAL COMPRA", "IMPORTE ICE", "TASAS", "SUBTOTAL", 
        "DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA", "CREDITO FISCAL"
    ]
    df_pendientes = df_pendientes[columnas_orden_tabla]
    
    edited_df = st.data_editor(
        df_pendientes,
        hide_index=True,
        column_config={
            "Guardar": st.column_config.CheckboxColumn("¿Aprobar?", help="Selecciona para enviar al reporte oficial", default=True),
            "IMPORTE TOTAL COMPRA": None,
            "IMPORTE ICE": None,
            "TASAS": None,
            "SUBTOTAL": None,
            "DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA": None,
            "CREDITO FISCAL": None
        },
        disabled=["Fecha", "Razón Social", "NIT", "Nro Factura", "IMPORTE BASE CF", "CUF / Autorización"],
        use_container_width=True
    )
    
    col_btn1, col_btn2 = st.columns([1, 1])
    with col_btn1:
        if st.button("✅ CONFIRMAR Y ENVIAR A LA NUBE", type="primary", use_container_width=True):
            df_aprobados = edited_df[edited_df["Guardar"] == True].drop(columns=["Guardar"])
            
            if not df_aprobados.empty:
                with st.spinner("Subiendo registros confirmados a Google Sheets..."):
                    guardar_historico(df_aprobados)
                    st.session_state.registros_sesion.extend(df_aprobados.to_dict('records'))
                st.success(f"¡Éxito! Se consolidaron {len(df_aprobados)} facturas en el sistema oficial.")
            else:
                st.warning("No aprobaste ninguna factura.")
                
            st.session_state.registros_pendientes = []
            st.rerun()
            
    with col_btn2:
        if st.button("🗑️ DESCARTAR TODAS", use_container_width=True):
            st.session_state.registros_pendientes = []
            st.rerun()

# --- REPORTE CONSOLIDADOS EN LA SESIÓN Y BOTONES DE DESCARGA ---
if st.session_state.registros_sesion:
    st.divider()
    st.write("### 📊 Registros Oficiales (Sesión Actual)")
    
    for i, reg in enumerate(st.session_state.registros_sesion):
        col_data, col_del = st.columns([12, 1])
        with col_data:
            monto_vista = reg.get('IMPORTE BASE CF', 0)
            st.markdown(f"""
            <div class='factura-card'>
                <span style='color: #741b28; font-weight: bold; font-size: 1.1em;'>{reg['Razón Social']}</span><br>
                <small>Factura: {reg['Nro Factura']} | Importe Base CF: {monto_vista} Bs. | NIT: {reg['NIT']}</small><br>
                <span class='cuf-text'>CUF: {reg['CUF / Autorización']}</span>
            </div>
            """, unsafe_allow_html=True)
        with col_del:
            st.write("")
            if st.button("✖", key=f"del_{i}", help="Eliminar permanentemente de la pantalla y de Google Sheets"):
                with st.spinner("Removiendo de la base de datos..."):
                    eliminar_de_historico_nube(reg['CUF / Autorización'])
                st.session_state.registros_sesion.pop(i)
                st.rerun()

    st.markdown("#### 📥 Panel de Descarga de Informes")
    df_historico_completo = cargar_historico()
    
    columnas_vista = ["Fecha", "Razón Social", "NIT", "Nro Factura", "IMPORTE BASE CF", "CUF / Autorización"]
    columnas_existentes_vista = [col for col in columnas_vista if col in df_historico_completo.columns]
    
    if not df_historico_completo.empty:
        st.dataframe(df_historico_completo[columnas_existentes_vista], use_container_width=True)
    
    # --- PROCESAMIENTO DEL INFORME TÉCNICO INTERNO ---
    columnas_validas_excel = [col for col in ORDEN_COLUMNAS_SISTEMA if col in df_historico_completo.columns]
    df_excel_final = df_historico_completo[columnas_validas_excel]
    
    buff_tecnico = BytesIO()
    with pd.ExcelWriter(buff_tecnico, engine='openpyxl') as w:
        df_excel_final.to_excel(w, index=False, sheet_name="Control Técnico")
        estilizar_hoja_excel(w.sheets["Control Técnico"])
        
    # --- PROCESAMIENTO DEL FORMATO DE IMPORTACIÓN SAP ---
    columnas_sap = [
        "Sociedad", "Código del Proveedor", "Fecha de Documento", "Fecha de Contabilización",
        "Número de Factura", "Número de Autorización", "Código de Control", "Condiciones de pago",
        "Razon Social", "NIT", "CUF", "MONTO RECIBO FACTURA (BOB)", "DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA"
    ]
    df_sap = pd.DataFrame(columns=columnas_sap)
    
    if not df_historico_completo.empty:
        num_filas = len(df_historico_completo)
        
        df_sap["Sociedad"] = ["BO01"] * num_filas
        df_sap["Código del Proveedor"] = [""] * num_filas
        df_sap["Fecha de Documento"] = df_historico_completo["Fecha"].values
        df_sap["Fecha de Contabilización"] = [""] * num_filas
        df_sap["Número de Factura"] = df_historico_completo["Nro Factura"].values
        df_sap["Número de Autorización"] = [""] * num_filas
        
        if "CODIGO CONTROL" in df_historico_completo.columns:
            df_sap["Código de Control"] = df_historico_completo["CODIGO CONTROL"].values
        else:
            df_sap["Código de Control"] = [""] * num_filas
            
        df_sap["Condiciones de pago"] = ["Z000"] * num_filas
        df_sap["Razon Social"] = df_historico_completo["Razón Social"].values
        df_sap["NIT"] = df_historico_completo["NIT"].values
        df_sap["CUF"] = df_historico_completo["CUF / Autorización"].values
        
        if "SUBTOTAL" in df_historico_completo.columns:
            df_sap["MONTO RECIBO FACTURA (BOB)"] = df_historico_completo["SUBTOTAL"].values
        else:
            df_sap["MONTO RECIBO FACTURA (BOB)"] = [0] * num_filas

        if "DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA" in df_historico_completo.columns:
            df_sap["DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA"] = df_historico_completo["DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA"].values
        else:
            df_sap["DESCUENTOS/BONIFICACIONES/REBAJAS SUJETAS AL IVA"] = [0] * num_filas
            
    df_sap = df_sap.fillna("")

    buff_sap = BytesIO()
    with pd.ExcelWriter(buff_sap, engine='openpyxl') as w:
        df_sap.to_excel(w, index=False, sheet_name="Layout SAP")
        estilizar_hoja_excel(w.sheets["Layout SAP"])
    
    # DESPLIEGUE GEOMÉTRICO DE DOS BOTONES PROFESIONALES
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button(
            label="📊 DESCARGAR INFORME TÉCNICO COMPLETO",
            data=buff_tecnico.getvalue(),
            file_name="Reporte_Nube_UNIVALLE.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col_dl2:
        st.download_button(
            label="⚙️ DESCARGAR FORMATO IMPORTACIÓN SAP",
            data=buff_sap.getvalue(),
            file_name="Carga_Masiva_SAP_UNIVALLE.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
else:
    if base_siat is None:
        st.info("📌 Sistema en espera. Por favor, vincule la base de datos diaria para iniciar el procesamiento.")
    elif not st.session_state.registros_pendientes:
        st.info("📌 Base operativa en la nube. Seleccione un método arriba (Enlaces o Búsqueda Manual) para encontrar registros.")

st.markdown("<br><p style='text-align: center; color: #741b28; opacity: 0.6;'>DEPARTAMENTO DE CONTABILIDAD | UNIVALLE S.A. © 2026</p>", unsafe_allow_html=True)
